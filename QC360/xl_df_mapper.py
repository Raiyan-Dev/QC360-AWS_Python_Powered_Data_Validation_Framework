import pandas as pd
from openpyxl import Workbook
from xlwings import Book as xlwings_book



def copy_missing_templates(existing_file_path, template_file_path, temp_list, logging):
    """
    Checks for missing sheets in the existing workbook compared to `temp_list`
    and copies them from the template workbook with formatting intact.
    
    :param existing_file_path: Path to the existing workbook.
    :param template_file_path: Path to the template workbook.
    :param temp_list: List of required sheet names.
    :param logging: Logger for logging information.
    """
    try:
        # Open the existing workbook and the template workbook using xlwings
        existing_wb = xlwings_book(existing_file_path)
        template_wb = xlwings_book(template_file_path)

        # Get the list of sheet names from the existing workbook
        existing_sheets = [sheet.name for sheet in existing_wb.sheets]

        # Find missing sheets by comparing temp_list with existing_sheets
        missing_sheets = [sheet for sheet in temp_list if sheet not in existing_sheets]

        # Log the missing sheets
        if missing_sheets:
            logging.info(f"Missing sheets in {existing_file_path}: {missing_sheets}")
            print(f"Missing sheets: {missing_sheets}")

        # Copy missing sheets from the template workbook
        for sheet in missing_sheets:
            try:
                # Access the template sheet
                template_sheet = template_wb.sheets[sheet]
                
                # Copy content (including formatting) from the template sheet
                template_sheet.api.Copy(After=existing_wb.sheets[-1].api)

                logging.info(f"Copied sheet '{sheet}' from template to {existing_file_path}")
                print(f"Copied sheet '{sheet}' successfully.")

            except Exception as e:
                logging.error(f"Error copying sheet '{sheet}': {str(e)}")
                print(f"Error copying sheet '{sheet}': {str(e)}")

        # Save and close the workbook
        existing_wb.save()
        existing_wb.close()

    except Exception as e:
        logging.error(f"Failed to check and copy missing sheets: {str(e)}")
        print(f"Failed to check and copy missing sheets: {str(e)}")



def map_df_to_excel(workbook: Workbook, template: str, df: pd.DataFrame, startRow: int, startCol: int):

    #Get the active worksheet
    ws = workbook[template]
    
    # Get the starting row and column
    start_row = startRow  
    start_col = startCol  
    
    # Write the DataFrame data starting from the specified row and column
    # Note:- startRow & startCol is given considering 'A1' Cell as [0,0]
    for row_num, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for col_num, value in enumerate(row, start=start_col+1):
            ws.cell(row=row_num, column=col_num, value=value)

    return


#This function takes in workbook and checks if the bu_tag, bu_id and qtr_nm in the unique_combinations and creates a new template name out of the template_name which is provided in the argument. Finally it creates that new sheet in the workbook and returns the new_template_name as the ouput.

# def create_dynamic_templates(workbook, template_name, unique_combinations, logging):
#     """
#     Creates dynamic templates by duplicating the given base template for each unique combination.

#     :param workbook: The openpyxl Workbook object.
#     :param template_name: Base template name to be duplicated.
#     :param unique_combinations: A dictionary containing BU_ID, QTR_NM, and BU_TAG combinations.
#     :param logging: Logger for logging information.
#     """
#     try:
#         if not isinstance(unique_combinations, dict):
#             raise ValueError("Expected unique_combinations to be a dictionary.")

#         bu_id = unique_combinations.get("BU_ID", "").strip().strip("'")
#         qtr_nm = unique_combinations.get("QTR_NM", "").strip().strip("'")
#         bu_tag = unique_combinations.get("BU_TAG", "").strip().strip("'")

#         if not bu_id or not qtr_nm:
#             logging.warning(f"FAILED : Incomplete data for combination. Skipping: {unique_combinations}")
#             print(f"FAILED : Incomplete data for combination. Skipping: {unique_combinations}")
#             return "Error"
#         # Generate the new template name
#         new_template_name = f"{template_name}_{bu_id}_{qtr_nm}"

#         if new_template_name in workbook.sheetnames:
#             logging.warning(f"Template '{new_template_name}' already exists. Skipping creation.")
#             print(f"Template '{new_template_name}' already exists. Skipping creation.")
            
#         else:
#             base_template = workbook[template_name]
#             new_sheet = workbook.copy_worksheet(base_template)
#             new_sheet.title = new_template_name

#             logging.info(f"Created new template: {new_template_name}")
#             print(f"Created new template: {new_template_name}")

#         return new_template_name

#     except Exception as e:
#         logging.error(f"Error creating template: {str(e)}")
#         print(f"Error creating template: {str(e)}")



def create_dynamic_templates_brd_nm(workbook, template_name, unique_combinations, brd_prod_id_dict, logging): 
    """
    Creates dynamic templates by duplicating the given base template for the matching BRD_NM.

    :param workbook: The openpyxl Workbook object.
    :param template_name: Base template name to be duplicated.
    :param unique_combinations: A dictionary containing MKT_ID, PROD_ID, and/or BRD_NM combinations.
    :param brd_prod_id_dict: A dictionary where keys are BRD_NM and values are corresponding PROD_IDs.
    :param logging: Logger for logging information.
    """
    try:
        if not isinstance(unique_combinations, dict):
            raise ValueError("Expected unique_combinations to be a dictionary.")
        if not isinstance(brd_prod_id_dict, dict):
            raise ValueError("Expected brd_prod_id_dict to be a dictionary.")

        # Extract PROD_ID, and BRD_NM from unique_combinations
        prod_id = unique_combinations.get("PROD_ID", "").strip().strip("'")
        brd_nm = unique_combinations.get("BRD_NM", "").strip().strip("'")

        # Check if BRD_NM is provided in unique_combinations
        if brd_nm:
            # Use BRD_NM directly
            matching_brd_nm = brd_nm
        elif prod_id:
            # Normalize PROD_ID values by removing extra quotes and spaces
            prod_ids = set(prod.strip("'").strip() for prod in prod_id.split(","))
            
            # Find the matching BRD_NM based on PROD_ID
            matching_brd_nm = next((
                brd_name
                for brd_name, prod_ids_str in brd_prod_id_dict.items()
                if prod_ids.intersection(set(prod.strip("'").strip() for prod in prod_ids_str.split(",")))
            ), None)
            
        else:
            logging.warning(f"FAILED : Neither BRD_NM nor PROD_ID found in unique_combinations. Skipping: {unique_combinations}")
            print(f"FAILED : Neither BRD_NM nor PROD_ID found in unique_combinations. Skipping: {unique_combinations}")
            return "Error"

        if not matching_brd_nm:
            logging.warning(f"FAILED : No matching BRD_NM found for PROD_ID: {prod_id}")
            print(f"FAILED : No matching BRD_NM found for PROD_ID: {prod_id}")
            return "Error"

        # Generate the new template name
        brd_nm_no_quotes = matching_brd_nm.strip("'")
        new_template_name = f"{template_name}_{brd_nm_no_quotes}" 

        if new_template_name not in workbook.sheetnames:
            # Duplicate the base template sheet
            base_template = workbook[template_name]
            new_sheet = workbook.copy_worksheet(base_template)
            new_sheet.title = new_template_name

            logging.info(f"Created new template: {new_template_name}")
            print(f"Created new template: {new_template_name}")

        return new_template_name

    except Exception as e:
        logging.error(f"Error creating template: {str(e)}")
        print(f"Error creating template: {str(e)}")
        return "Error"
