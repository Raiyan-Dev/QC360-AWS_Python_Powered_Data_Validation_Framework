#Script Summary

import pandas as pd
from collections import defaultdict
from openpyxl import load_workbook
from Query import run_query_parallel
from pathlib import Path
from io import BytesIO
from xl_df_mapper import map_df_to_excel, copy_missing_templates, create_dynamic_templates_brd_nm



def OVR(queries_df_OVR_intm, config_df, filter_dict, mapping_list_df, inpt_ctrl_cfg_df, unique_bu_qtr_df, process, env_dict, env, logging):

    # Making a Combo to loop for different Workbooks (File Name) and directory (Path)
    OVR_combo = queries_df_OVR_intm[['Path', 'File Name']].drop_duplicates().reset_index(drop=True)
    
    # Fetch the qtr_nm, bu_id and bu_tag from the 'unique_bu_qtr_df' argument
    qtr_nm = unique_bu_qtr_df['QTR_NM'].iloc[0]
    bu_id = unique_bu_qtr_df['BU_ID'].iloc[0]
    bu_tag = unique_bu_qtr_df['BU_TAG'].iloc[0]
    
    #Mapping type dictionary
    mapping_combo_dict = {0: None, 
                    1: 'ALL', 
                    2: ['MKT_ID', 'PROD_ID'],
                    3: ['COMPLETE_TEAM_NM', 'MKT_ID', 'PROD_ID'],
                    4: ['MKT_ID', 'BRD_NM'],
                    5: ['COMPLETE_TEAM_NM', 'MKT_ID', 'BRD_NM'],
                    6: ['BRD_NM'],
                    7: ['TEAM_NM'],
                    8: ['BRD_FAM_DS', 'TEAM_NM_LIKE'],
                    9: ['MIRROR_TEAM_ATNT', 'MKT_ID_ATNT'],
                    10: ['MKT_ID', 'PROD_ID', 'BRD_NM'],
                    11: ['COMPLETE_TEAM_NM', 'MKT_ID', 'BRD_NM', 'PROD_ID'],
                    12: ['LATEST_MONTH'],
                    13: ['BRD_NM', 'TEAM_NM', 'MKT_NM'],
                    14: ['MKT_NM'],
                    15: ['MKT_NM', 'BRD_NM', 'PARENT_TEAM'],
                    16: ['MKT_NM', 'BRD_NM', 'PARENT_TEAM', 'COMPLETE_TEAM_NM'],
                    17: ['SLS_MTRC', 'BRD_NM'],
                    18: ['BS_SPEC_TEAM_NM', 'BRD_NM'],
                    19: ['BRD_NM', 'TEAM_NM'],
                    20: ['BRD_FAM_NM', 'COMPLETE_TEAM_NM', 'PARENT_TEAM', 'LATEST_MONTH', 'SLS_MTRC'],
                    21: ['COL_SLS_MTRC', 'BRD_NM'],
                    22: ['COMPLETE_TEAM_NM', 'MKT_NM'],
                    23: ['COMPLETE_TEAM_NM', 'MKT_NM', 'SIP_ORALS_BRD_NM'],
                    24: ['COMPLETE_TEAM_NM', 'MKT_NM', 'SIP_INFUSIONS_BRD_NM'],
                    25: ['LVL_ID_1', 'SIP_INFUSIONS_BRD_NM', 'COMPLETE_TEAM_NM', 'BRD_LIKE', 'PARENT_TEAM', 'MKT_NM'],
                    26: ['LVL_ID_1', 'SIP_ORALS_BRD_NM', 'COMPLETE_TEAM_NM', 'BRD_LIKE', 'PARENT_TEAM', 'MKT_NM'],
                    27: ['COMPLETE_TEAM_NM', 'PARENT_TEAM', 'LVL_2_TEAM', 'LVL_5_TEAM', 'MKT_NM', 'CHNL_ID'],
                    28: ['COMPLETE_TEAM_NM', 'PARENT_TEAM', 'LVL_2_TEAM', 'LVL_5_TEAM', 'MKT_NM', 'CHNL_ID','SIP_ORALS_BRD_NM'],
                    29: ['COMPLETE_TEAM_NM', 'PARENT_TEAM', 'LVL_2_TEAM', 'LVL_5_TEAM', 'MKT_NM', 'CHNL_ID','SIP_INFUSIONS_BRD_NM'],
                    30: ['COMPLETE_TEAM_NM', 'LVL_2_TEAM', 'LVL_5_TEAM', 'MKT_NM', 'SIP_ORALS_BRD_NM'],
                    31: ['COMPLETE_TEAM_NM', 'LVL_2_TEAM', 'LVL_5_TEAM', 'MKT_NM', 'SIP_INFUSIONS_BRD_NM'],
                    32: ['COMPLETE_TEAM_NM', 'LVL_2_TEAM', 'LVL_5_TEAM', 'MKT_NM'],
                    33: ['COMPLETE_TEAM_NM', 'MKT_NM', 'SIP_INFUSIONS_BRD_NM', 'LVL_ID_1'],
                    34: ['COMPLETE_TEAM_NM', 'MKT_NM', 'SIP_ORALS_BRD_NM', 'LVL_ID_1'],
                    35: ['COMPLETE_TEAM_NM', 'MKT_NM', 'LVL_ID_1'],
                    36: ['COMPLETE_TEAM_NM', 'MKT_NM', 'BRD_NM'],
                    37: ['BRD_NM', 'MKT_NM', 'LVL_5_TEAM'],
                    38: ['COMPLETE_TEAM_NM', 'PARENT_TEAM', 'MKT_NM', 'LVL_ID_1', 'SIP_ORALS_BRD_NM'],
                    39: ['COMPLETE_TEAM_NM', 'PARENT_TEAM', 'MKT_NM', 'LVL_ID_1', 'SIP_INFUSIONS_BRD_NM'],
                    40: ['LATEST_MONTH', 'CLR_BS_SUM_OF_MONTH_LATEST'],
                    41: ['LATEST_MONTH', 'BRD_NM'],
                    42: ['BRD_NM', 'CLR_BS_SUM_OF_MONTH_LATEST'],
                    43: ['COMPLETE_TEAM_NM']
                    }
            
    # Get the column indices of 'QTR_NM' and 'Process'. Get all columns b/w this strt_col & end_col
    strt_col = mapping_list_df.columns.get_loc('QTR_NM')
    end_col = mapping_list_df.columns.get_loc('Process')
    # Create a new DataFrame with the desired columns for mapping type NON-ZERO to loop on.
    looping_df = mapping_list_df.iloc[:, strt_col+1:end_col]     
    
    for j in range(0, len(OVR_combo)):
        try:
            file_path = str(OVR_combo['Path'][j]).strip()
            file_name_tmp = str(OVR_combo['File Name'][j]).strip()

            file_name = "{name}_{bu}_{qtr}".format(
                name=str(OVR_combo['File Name'][j]).strip(),
                bu=str(bu_id).strip().replace("'", ''),
                qtr=str(qtr_nm).strip().replace("'", '')
            )

            file_full_path = f"{file_path}\\{file_name}.xlsx"
            
            template_names = queries_df_OVR_intm.loc[
                (queries_df_OVR_intm['Path'] == file_path) & (queries_df_OVR_intm['File Name'] == file_name_tmp)].reset_index(
                drop=True)[['Template Name']].drop_duplicates().reset_index(drop=True)

            temp_list = [x for x in template_names['Template Name']]
            
            if Path(file_full_path).is_file():
                
                print('File ' + file_full_path + ' already present. Using this file for Execution \n')
                logging.info('File ' + file_full_path + ' already present. Using this file for Execution \n')

                # Check and copy missing sheets using xlwings
                copy_missing_templates(
                    existing_file_path=file_full_path,
                    template_file_path="reporting_config.xlsm",
                    temp_list=temp_list,
                    logging=logging
                )
                
                book = load_workbook(file_full_path, data_only=False)
                
            else:
                #load workbook directly from reporting_config itself
                book = load_workbook('reporting_config.xlsm', data_only=False)
                print('Using New Templates for File ' + file_full_path)
                logging.info('Using New Templates for File ' + file_full_path)
    
                for name in book.sheetnames:
                    if name not in temp_list:
                        del book[name]


            queries_df_OVR = queries_df_OVR_intm.loc[
                (queries_df_OVR_intm['Path'] == file_path) & (queries_df_OVR_intm['File Name'] == file_name_tmp)].reset_index(drop=True)
            
            queriesList = []

            for i in range(0, len(queries_df_OVR)):
                query = ""

                # Loop through the columns 'Query_1', 'Query_2', 'Query_3', 'Query_4' to concatenate contents
                for col in ['Query_1', 'Query_2', 'Query_3', 'Query_4']:
                    # Get the current query part from the dataframe
                    query_part = queries_df_OVR[col][i]
                    
                    # Check if the query part is valid (not NaN or blank)
                    if isinstance(query_part, str) and query_part.strip() != '':
                        # Append the non-empty query part with a space
                        query += query_part.strip() + " "
                    elif isinstance(query_part, float) and not pd.isna(query_part):
                        query += str(query_part).strip() + " "

                # Remove any trailing spaces & unwanted characters (\n, \t etc)
                query = query.replace('\n', ' ').replace('\t', ' ').strip()

                # Get the mapping type for current iteration and lookup the corresponding value in mapping_combo_dict
                mapping_type = queries_df_OVR['Mapping_type'][i]
                mapping_value = mapping_combo_dict.get(mapping_type)

                # Create a base query for the current BU_ID and QTR_NM combination
                query_k = query.replace("$$BU_ID".lower(), f"{str(bu_id)}")
                query_k = query_k.replace("$$QTR_NM".lower(), f"{str(qtr_nm)}")
                query_k = query_k.replace("$$BU_TAG".lower(), f"{str(bu_tag)}")
                
                # Replace NON-LOOPING $$column_name variables as well (All the columns after the Process_Active_Flag)
                start_col_idx = unique_bu_qtr_df.columns.get_loc('Process_Active_Flag') + 1

                for col in unique_bu_qtr_df.columns[start_col_idx:]:
                    placeholder = f"$${col.lower()}"
                    if placeholder in query_k:
                        query_k = query_k.replace(placeholder, f"{str(unique_bu_qtr_df[col].values[0])}")
                
                # Check for DATA_DT columns like DATA_DT_1, DATA_DT_2, etc. and handle NaN
                for num in range(1, 8):
                    table_col = f'TABLE_{num}'
                    data_dt_col = f'DATA_DT_{num}'
                    
                    full_table_name = str(queries_df_OVR[table_col].iloc[i]). strip()
                    # Check if TABLE_num is NaN; if so, break out of the loop
                    if pd.isna(full_table_name):
                        continue
                    if isinstance(full_table_name, str):
                        table_name = full_table_name.split('.')[-1]

                        # If table name starts with 'az_jbrm' and ends with '_obu', remove '_obu'
                        if table_name.startswith('az_jbrm') and table_name.endswith('_obu'):
                            table_name = table_name[:-4]  # Removes the last 4 characters ('_obu')
                            
                    else:
                        table_name = None  
                    
                    # Check if the DATA_DT is missing (NaN) and table_name is present
                    if pd.isna(queries_df_OVR[data_dt_col].iloc[i]) and table_name:
                        # Fetch the corresponding data_dt from inpt_ctrl_cfg_df matching the table name, qtr_nm and bu_tag
                        
                        matching_rows = inpt_ctrl_cfg_df[
                            (inpt_ctrl_cfg_df['tbl_nm'] == table_name) &
                            (inpt_ctrl_cfg_df['qtr_nm'] == qtr_nm.strip().replace("'", '')) &
                            (inpt_ctrl_cfg_df['bu_tag'] == bu_tag.strip().replace("'", ''))
                        ]
                        
                        data_dt_value = matching_rows['attr_val_1'].values
                        
                        # Check if a value was found, otherwise set a default
                        if data_dt_value.size > 0:
                            data_dt_value = f"'{data_dt_value[0]}'"
                        else:
                            data_dt_value = "''"
                    else:
                        # Use the existing value if present
                        data_dt_value = str(queries_df_OVR[data_dt_col].iloc[i])
                
                    # Replace the $$DATA_DT_X placeholder with the found value
                    query_k = query_k.replace(f"$$DATA_DT_{num}".lower(), f"{data_dt_value}")
                
                # Now proceed with specific replacements based on the `mapping_value` in `mapping_combo_dict`
                if mapping_value is not None:
                    # If mapping_value is 'ALL', loop through All columns and rows in looping_df
                    if mapping_value == 'ALL':
                        for j in range(len(looping_df)):
                            query_j = query_k
                            looping_df = looping_df.fillna("''")

                            # Replace $$column_name variables with values from the current row of looping_df
                            for col in looping_df.columns:
                                placeholder = f"$${col.lower()}"
                                if placeholder in query_j:
                                    query_j = query_j.replace(placeholder, f"{str(looping_df[col].iloc[j])}")

                            # Additional replacements with filter_dict values
                            for column in queries_df_OVR:
                                query_j = query_j.replace("$$env", str(env_dict.get(env)))
                                if column == 'Process':
                                    if queries_df_OVR[column][i] == 'Goals':
                                        query_j = query_j.replace(str(filter_dict.get(column)), '_goals_s')
                                    else:
                                        query_j = query_j.replace(str(filter_dict.get(column)), '')

                                if filter_dict.get(column) is not None and pd.notna(queries_df_OVR[column][i]):
                                    query_j = query_j.replace(str(filter_dict.get(column)), str(queries_df_OVR[column][i]))

                            queriesList.append({
                                'QC_No': queries_df_OVR["QC No."][i],
                                'QC_Desc': queries_df_OVR["QC Description"][i],
                                'Query': query_j,
                                'Template': queries_df_OVR["Template Name"][i],
                                'StartRow': int(queries_df_OVR["Start Row"][i]),
                                'StartCol': int(queries_df_OVR["Start Column"][i])
                            })  

                    # If mapping_value is a list (e.g., [TEAM_NM, MKT_NM, CHNL_ID])
                    elif isinstance(mapping_value, list):
                        unique_combinations_df = looping_df[mapping_value].drop_duplicates().dropna(how='all').reset_index(drop=True)
                        
                        unique_combinations_df = unique_combinations_df.fillna("''")
                        
                        brd_prod_id_df = looping_df[['BRD_NM', 'PROD_ID']].drop_duplicates().dropna(how='all').reset_index(drop=True)
                        brd_prod_id_dict = brd_prod_id_df.set_index('BRD_NM')['PROD_ID'].to_dict()

                        for j in range(len(unique_combinations_df)):

                            if process.lower() == 'ovr':

                                tmplt_nm = create_dynamic_templates_brd_nm(
                                    book, 
                                    queries_df_OVR["Template Name"][i],
                                    unique_combinations_df.iloc[j].to_dict(),
                                    brd_prod_id_dict,
                                    logging
                                )

                                # Start with query that has BU_ID, QTR_NM and BU_TAG already replaced
                                query_j = query_k  
                                
                                # Replace $$column_name variables based on unique combinations
                                for col in mapping_value:
                                    placeholder = f"$${col.lower()}"
                                    if placeholder in query_j:
                                        query_j = query_j.replace(placeholder, f"{str(unique_combinations_df[col].iloc[j])}")
                                
                                # Additional replacements with filter_dict values
                                for column in queries_df_OVR:
                                    query_j = query_j.replace("$$env", str(env_dict.get(env)))
                                    if column == 'Process':
                                        if queries_df_OVR[column][i] == 'Goals':
                                            query_j = query_j.replace(str(filter_dict.get(column)), '_goals_s')
                                        else:
                                            query_j = query_j.replace(str(filter_dict.get(column)), '')

                                    if filter_dict.get(column) is not None and pd.notna(queries_df_OVR[column][i]):
                                        query_j = query_j.replace(str(filter_dict.get(column)), str(queries_df_OVR[column][i]))

                                queriesList.append({
                                    'QC_No': queries_df_OVR["QC No."][i],
                                    'QC_Desc': queries_df_OVR["QC Description"][i],
                                    'Query': query_j,
                                    'Template': tmplt_nm,
                                    'StartRow': int(queries_df_OVR["Start Row"][i]),
                                    'StartCol': int(queries_df_OVR["Start Column"][i])
                                })

                            else:
                                # Start with query that has BU_ID, QTR_NM and BU_TAG already replaced
                                query_j = query_k  
                                
                                # Replace $$column_name variables based on unique combinations
                                for col in mapping_value:
                                    placeholder = f"$${col.lower()}"
                                    if placeholder in query_j:
                                        query_j = query_j.replace(placeholder, f"{str(unique_combinations_df[col].iloc[j])}")
                                
                                # Additional replacements with filter_dict values
                                for column in queries_df_OVR:
                                    query_j = query_j.replace("$$env", str(env_dict.get(env)))
                                    if column == 'Process':
                                        if queries_df_OVR[column][i] == 'Goals':
                                            query_j = query_j.replace(str(filter_dict.get(column)), '_goals_s')
                                        else:
                                            query_j = query_j.replace(str(filter_dict.get(column)), '')

                                    if filter_dict.get(column) is not None and pd.notna(queries_df_OVR[column][i]):
                                        query_j = query_j.replace(str(filter_dict.get(column)), str(queries_df_OVR[column][i]))

                                queriesList.append({
                                    'QC_No': queries_df_OVR["QC No."][i],
                                    'QC_Desc': queries_df_OVR["QC Description"][i],
                                    'Query': query_j,
                                    'Template': queries_df_OVR["Template Name"][i],
                                    'StartRow': int(queries_df_OVR["Start Row"][i]),
                                    'StartCol': int(queries_df_OVR["Start Column"][i])
                                })           
                                
                # If mapping_value is None, perform direct replacements as usual
                else:
                    query_k = query_k  # Already has BU_ID, QTR_NM and BU_TAG replaced
                    
                    for column in queries_df_OVR:
                        query_k = query_k.replace("$$env", str(env_dict.get(env)))
                        if column == 'Process':
                            if queries_df_OVR[column][i] == 'Goals':
                                query_k = query_k.replace(str(filter_dict.get(column)), '_goals_s')
                            else:
                                query_k = query_k.replace(str(filter_dict.get(column)), '')

                        if filter_dict.get(column) is not None and pd.notna(queries_df_OVR[column][i]):
                            query_k = query_k.replace(str(filter_dict.get(column)), str(queries_df_OVR[column][i]))

                    queriesList.append({
                        'QC_No': queries_df_OVR["QC No."][i],
                        'QC_Desc': queries_df_OVR["QC Description"][i],
                        'Query': query_k,
                        'Template': queries_df_OVR["Template Name"][i],
                        'StartRow': int(queries_df_OVR["Start Row"][i]),
                        'StartCol': int(queries_df_OVR["Start Column"][i])
                    })

            try:
                response_list = run_query_parallel(queriesList, config_df, logging)
            except RuntimeError as runtimeError:
                raise runtimeError
            except Exception as QueryRunException:
                print("Some Error in running the queries on Athena\n")
                print("ERROR MESSAGE: ")
                print(str(QueryRunException))
                return

        

            #########################                                             #########################
            ######################### Writing the Result of the Queries to Excel  #########################
            #########################                                             #########################
            

            # A dictionary to store the queries by grouping them based on their QC_No & Template
            grouped_results = defaultdict(list)

            for response in response_list:
                grouped_results[(response['QC_No'], response['Template'])].append(response)

            # Iterate over grouped results
            for (qc_no, template), responses in grouped_results.items():
                print(
                    "------------------------------------------------------------------------------------------------------------------------------------------")
                logging.info(
                    "------------------------------------------------------------------------------------------------------------------------------------------")
                try:
                    # Combine all Result dataframes for the same QC_No
                    combined_df = pd.DataFrame()  # Initialize an empty dataframe
                    for response in responses:
                        if response['Status'] == 'SUCCEEDED':
                            # Check if the Result dataframe is not empty
                            if not response['Result'].empty:
                                # Concatenate the Result dataframe
                                combined_df = pd.concat([combined_df, response['Result']], ignore_index=True)
                        elif response['Status'] == 'FAILED':
                            # Append a placeholder row indicating failure
                            combined_df = pd.concat([combined_df, pd.DataFrame(['e'], columns=['col1'])], ignore_index=True)

                    # Check if the combined dataframe is empty (handle the edge case where all responses fail or have no data)
                    if combined_df.empty:
                        combined_df = pd.DataFrame()  # Empty dataframe to write nothing in Excel

                    primary_response = responses[0]

                    # Write the combined dataframe to Excel
                    map_df_to_excel(
                        workbook=book,
                        template=primary_response['Template'].strip(),
                        df=combined_df,
                        startRow=int(primary_response["StartRow"]),
                        startCol=int(primary_response["StartCol"])
                    )
                    qc_no = int(qc_no)
                    print(f"Successfully Completed QC No: {qc_no} with {len(responses)} queries combined.\n")
                    logging.info(f"Successfully Completed QC No: {qc_no} with {len(responses)} queries combined.\n")

                    # Log all query executions for this QC_No
                    for response in responses:
                        print(
                            "-------------------------------------------------------------------------------------------------------------------------------")
                        logging.info(
                            "------------------------------------------------------------------------------------------------------------------------------------------")
                        if response['Status'] == 'SUCCEEDED':
                            print(f"Successfully Completed QC No: {qc_no} : {response['QC_Desc']}:-\n\nWritten on template - {template}\n")
                            logging.info(f"Successfully Completed QC No: {qc_no} : {response['QC_Desc']}:-\n\nWritten on template - {template}\n")
                            
                            print(f"Query which was executed - {response['Query']}\n")
                            logging.info(f"Query which was executed - {response['Query']}\n")
                            
                            print(f"Execution Id: {response['ExecutionId']}\n")
                            logging.info(f"Execution Id: {response['ExecutionId']}\n")
                        
                        elif response['Status'] == 'FAILED':
                            print(f"Failed to execute QC No: {qc_no} : {response['QC_Desc']}:-\n")
                            logging.info(f"Failed to execute QC No: {qc_no} : {response['QC_Desc']}:-\n")

                            print(f"Query which was failed to execute - {response['Query']}\n")
                            logging.info(f"Query which was failed to execute - {response['Query']}\n")
                            
                            print(f"ERROR MESSAGE:- {response['Error']}\n")
                            logging.info(f"ERROR MESSAGE:- {response['Error']}\n")
                            
                            if response['ExecutionId'] is None:
                                print("Execution Id: null\n\n")
                                logging.info("Execution Id: null\n\n")
                            else:
                                print(f"Execution Id: {response['ExecutionId']}\n")
                                logging.info(f"Execution Id: {response['ExecutionId']}\n")
                
                except Exception as sheetWriteException:
                    # Handle exceptions during Excel writing
                    print(f"Query {qc_no} {response['QC_Desc']}, could not be executed due to an unknown ERROR:-\n")
                    logging.info(f"Query {qc_no} {response['QC_Desc']}, could not be executed due to an unknown ERROR:-\n")
                    print(f"Query: {response['Query']}\n")
                    logging.info(f"Query: {response['Query']}\n")
                    print(f"ERROR MESSAGE:- {sheetWriteException}\n\n")
                    logging.info(f"ERROR MESSAGE:- {sheetWriteException}\n\n")
            
            # Create an in-memory binary stream to hold the Excel workbook's data and save it
            output_stream = BytesIO()
            book.save(output_stream)

            # Move the pointer of the binary stream back to the beginning
            output_stream.seek(0)
            # Now save the workbook directly to the specified file path + file name on disk
            book.save(file_full_path)

        except Exception as finalException:
            raise finalException